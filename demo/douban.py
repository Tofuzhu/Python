import re
import time

import requests
from openpyxl import Workbook

import jieba

def getTop250():
    FilmName = []
    FilmRating = []
    FilmPeople = []
    FilmQuote = []
    FilmId = []

    for urlnum in range(0, 250, 25):
        r = requests.get(r"https://movie.douban.com/top250?start=" + str(urlnum))
        w = open("1.txt", 'w+', encoding="utf-8")
        w.write(r.text)
        w.write(str(r.json))
        w.close()
        # 获取本页的电影名称
        pattern_Name = re.compile(r"<span class=\"title\">(.*?)</span>")
        m = pattern_Name.findall(r.text)
        film_Name = m[:]
        for i in m:
            if "nbsp" in i:
                film_Name.remove(i)
        print(film_Name)
        FilmName += film_Name

        # 获取本页的评分
        pattern_Rating = re.compile((r"<span class=\"rating_num\" property=\"v:average\">(.*?)</span>"))
        film_Rating = pattern_Rating.findall(r.text)
        film_Ratings = list(map(float, film_Rating))
        print(film_Ratings)
        FilmRating += film_Ratings

        # 获取观看人数
        pattern_People = re.compile((r"<span>(.*?)人评价</span>"))
        film_People = pattern_People.findall(r.text)
        film_Peoples = list(map(int, film_People))
        print(film_Peoples)
        FilmPeople += film_Peoples

        # 获取影片短评
        pattern_Quote = re.compile((r"<span class=\"inq\">(.*?)</span>"))
        film_Quote = pattern_Quote.findall(r.text)
        print(film_Quote)
        FilmQuote += film_Quote

        # 获取豆瓣id
        pattern_id = re.compile(r"<a href=\"https://movie.douban.com/subject/(.*?)/\" class=\"\">")
        film_Id = pattern_id.findall(r.text)
        print(film_Id)
        FilmId += film_Id

        time.sleep(1)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名称"
    ws["B1"] = "评分"
    ws["C1"] = "参评人数"
    ws["D1"] = "简评"
    ws["E1"] = "豆瓣ID"

    for id in range(0, 250):
        ws.cell(row=id + 2, column=1).value = FilmName[id]
        ws.cell(row=id + 2, column=2).value = FilmRating[id]
        ws.cell(row=id + 2, column=3).value = FilmPeople[id]
        ws.cell(row=id + 2, column=4).value = FilmQuote[id]
        ws.cell(row=id + 2, column=5).value = FilmId[id]

    wb.save(r"moviedata.xlsx")


def getMoviecomments(movieId):
    commentAllpage = []
    for commentNum in range(0, 1000, 20):
        r = requests.get(r"https://movie.douban.com/subject/" + str(movieId) + "/comments?start=" + str(
            commentNum) + "&limit=20&sort=new_score&status=P")
        w = open("1.txt", 'w+', encoding="utf-8")
        w.write(r.text)
        w.write(str(r.json))
        w.close()
        patternComment = re.compile(r"<span class=\"short\">(.*?)</span>")
        commentOnepage = patternComment.findall(r.text)
        # print(commentOnepage)
        commentAllpage += commentOnepage
        time.sleep(0.5)
    print(commentAllpage)


# def wordCloud(content):
#     str(content).replace("")
#     sd



if __name__ == "__main__":
    # getTop250()
    getMoviecomments(19971676)
