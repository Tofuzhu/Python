import openpyxl as op

# from openpyxl.utils import get_column_letter
path = r"C:\Users\zzy\Desktop\\"
wb = op.load_workbook(path+'RD45,RD46,RD51,RD51_20190530_v2.xlsx')
print(wb.get_sheet_names())
if wb.get_sheet_by_name('asd'):
    ws4 = wb.get_sheet_by_name('asd')
else:
    ws4 = wb.create_sheet("asd")
print(wb.get_sheet_names())
ws4['a1'] = 'asdsdfasdfasdf'
ws4['b2'] = 55
ws4.cell(3,4).value = 442323
sheet3 = wb.get_sheet_by_name("Sheet3")
sheet3['c3'].value = 38
sheet3.append(range(40))
print("Worksheet range(s):", wb.get_named_ranges())
print("Worksheet name(s):", wb.get_sheet_names())
ws1 = wb.get_sheet_by_name('RD45')
print(ws1.max_row)
print(ws1.max_column)
data = {}
for row in range(1,ws1.max_row):
    for column in range(1,ws1.max_column):
        templist=[]
        pid = row
        templist.append(ws1.cell(row=row,column=column).value)
        data[pid] = templist
print(data)
wb.save(path+'RD45,RD46,RD51,RD51_20190530_v2.xlsx')


#
# ws1 = wb.active
# ws1.title = "range names"
#
# for row in range(1,40):
#     ws1.append(range(600))
#
# ws2 = wb.create_sheet(title='Pi')
#
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title='Data')
#
# for row in range(10,20):
#     for col in range(27,54):
#         _ = ws3.cell(column=col,row=row,value='{0}'.format(get_column_letter(col)))
# print(ws3['AA10'].value)
#
# wb.save(filename = "pyexcel.xlsx")